def process_files(prior_csv, current_csv, sales_csv, delayed_csv, client_device_csv, selected_month, selected_year, save_path):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from datetime import datetime
    from tkinter import messagebox
    import os

    # Load the CSV files into DataFrames with specified dtype to avoid mixed types
    df_prior = pd.read_csv(prior_csv, dtype=str)
    df_current = pd.read_csv(current_csv, dtype=str)
    df_sales = pd.read_csv(sales_csv, dtype=str)
    df_delayed = pd.read_csv(delayed_csv, dtype=str)
    df_client_device = pd.read_csv(client_device_csv, dtype=str, skiprows=1)  # Skip the first line

    # Strip leading/trailing spaces from column names
    df_prior.columns = df_prior.columns.str.strip()
    df_current.columns = df_current.columns.str.strip()
    df_sales.columns = df_sales.columns.str.strip()
    df_delayed.columns = df_delayed.columns.str.strip()
    df_client_device.columns = df_client_device.columns.str.strip()

    # Rename "CustomerCode" to "SabreCode"
    df_prior.rename(columns={'CustomerCode': 'SabreCode'}, inplace=True)
    df_current.rename(columns={'CustomerCode': 'SabreCode'}, inplace=True)
    df_sales.rename(columns={'CustomerCode': 'SabreCode'}, inplace=True)
    df_delayed.rename(columns={'CustomerCode': 'SabreCode'}, inplace=True)
    df_client_device.rename(columns={'CustomerCode': 'SabreCode'}, inplace=True)

    # Convert Quantity and TotalDue columns to numeric
    df_current['Quantity'] = pd.to_numeric(df_current['Quantity'])
    df_current['TotalDue'] = pd.to_numeric(df_current['TotalDue'])
    df_prior['Quantity'] = pd.to_numeric(df_prior['Quantity'])

    # Consolidate Current month data
    df_current_consolidated = df_current.groupby(
        ['SabreCode', 'BranchName', 'ItemCode', 'BillingAmount']
    ).agg(
        Quantity_Current=('Quantity', 'sum'),
        TotalDue=('TotalDue', 'sum')
    ).reset_index()

    # Consolidate Prior month data
    df_prior_consolidated = df_prior.groupby(
        ['SabreCode', 'BranchName', 'ItemCode', 'BillingAmount']
    ).agg(
        Quantity_Previous=('Quantity', 'sum')
    ).reset_index()

    # Convert Quantity and BillingAmount columns to numeric
    df_current_consolidated['Quantity_Current'] = pd.to_numeric(df_current_consolidated['Quantity_Current'])
    df_prior_consolidated['Quantity_Previous'] = pd.to_numeric(df_prior_consolidated['Quantity_Previous'])
    df_current_consolidated['BillingAmount'] = pd.to_numeric(df_current_consolidated['BillingAmount'])
    df_prior_consolidated['BillingAmount'] = pd.to_numeric(df_prior_consolidated['BillingAmount'])

    # Merge the DataFrames on common columns
    df_merged = pd.merge(
        df_current_consolidated,
        df_prior_consolidated,
        on=['SabreCode', 'BranchName', 'ItemCode', 'BillingAmount'],
        how='left'
    )

    # Calculate the 'Dif' column
    df_merged['Dif'] = df_merged['Quantity_Current'] - df_merged['Quantity_Previous'].fillna(0)

    # Convert 'Qty' column to numeric to ensure proper summation
    df_sales['Qty'] = pd.to_numeric(df_sales['Qty'], errors='coerce').fillna(0)

    # Consolidate Sales data
    df_sales_consolidated = df_sales.groupby(
        ['SabreCode', 'TechtoolCode']
    ).agg(
        Sales_Qty=('Qty', 'sum')
    ).reset_index()

    # Clean 'Sales_Qty' column to remove or replace invalid values
    df_sales_consolidated['Sales_Qty'] = pd.to_numeric(df_sales_consolidated['Sales_Qty'], errors='coerce').fillna(0)

    # Explicitly cast Sales_Qty to float
    df_sales_consolidated['Sales_Qty'] = df_sales_consolidated['Sales_Qty'].astype(float)

    # Merge Sales data into the merged DataFrame
    df_merged['Sales'] = 0  # Initialize the Sales column
    df_merged['Sales'] = df_merged['Sales'].astype(float)  # Explicitly cast to float
    for idx, row in df_merged.iterrows():
        if row['Dif'] > 0:
            sales_data = df_sales_consolidated[
                (df_sales_consolidated['SabreCode'] == row['SabreCode']) &
                (df_sales_consolidated['TechtoolCode'] == row['ItemCode'])  # Assuming 'TechtoolCode' maps to 'ItemCode'
            ]
            if not sales_data.empty:
                df_merged.at[idx, 'Sales'] = sales_data['Sales_Qty'].sum()
        elif row['Dif'] == 0:
            # Filter for duplicate 'ItemCode' in df_merged and use the last occurrence
            duplicate_itemcodes = df_merged[
                (df_merged['SabreCode'] == row['SabreCode']) &
                (df_merged['ItemCode'] == row['ItemCode'])
            ]
            if not duplicate_itemcodes.empty:
                last_occurrence_idx = duplicate_itemcodes.index[-1]
                sales_data = df_sales_consolidated[
                    (df_sales_consolidated['SabreCode'] == row['SabreCode']) &
                    (df_sales_consolidated['TechtoolCode'] == row['ItemCode'])  # Assuming 'TechtoolCode' maps to 'ItemCode'
                ]
                if not sales_data.empty:
                    df_merged.at[last_occurrence_idx, 'Sales'] = sales_data['Sales_Qty'].sum()

    # Update 'Sales' column to 'No Sales found' where 'Dif' > 0 and 'Sales' is 0
    df_merged['Sales'] = df_merged.apply(lambda row: 'No Sales found' if row['Dif'] > 0 and row['Sales'] == 0 else row['Sales'], axis=1)

    # Add 'Delayed Billing' column
    delayed_codes = df_delayed['SabreCode'].unique()
    df_merged['Delayed Billing'] = df_merged['SabreCode'].apply(lambda x: 'Delayed Billing' if x in delayed_codes else '')

    # Add 'Actual added Current month' column
    selected_date = f"{selected_year}-{int(selected_month):02d}"
    
    # Ensure 'ManufactureDate' column exists (case-insensitive check)
    df_client_device.columns = df_client_device.columns.str.strip()  # Ensure no leading/trailing spaces
    manufacture_date_col = next((col for col in df_client_device.columns if col.lower() == 'manufacturedate'), None)
    
    if manufacture_date_col:
        df_client_device[manufacture_date_col] = pd.to_datetime(df_client_device[manufacture_date_col]).dt.to_period('M')
    else:
        messagebox.showerror("Error", "'ManufactureDate' column not found in Client Device CSV.")
        return
    
    def calculate_actual_added(row):
        if row['Dif'] > 0:
            actual_added = df_client_device[
                (df_client_device['SabreCode'] == row['SabreCode']) &
                (df_client_device['ItemCode'] == row['ItemCode']) &
                (df_client_device[manufacture_date_col] == selected_date)
            ]['ItemCode'].count()
            return actual_added
        return 0

    df_merged['Actual added Current month'] = df_merged.apply(calculate_actual_added, axis=1)

    # Replace 0 values with '-' in 'Dif', 'Sales', and 'Actual added Current month' columns
    df_merged['Dif'] = df_merged['Dif'].map(lambda x: '-' if x == 0 else x)
    df_merged['Sales'] = df_merged['Sales'].map(lambda x: '-' if x == 0 and x != 'No Sales found' else x)
    df_merged['Actual added Current month'] = df_merged['Actual added Current month'].map(lambda x: '-' if x == 0 else x)

    # Add 'Notes' column and populate it based on conditions
    def generate_notes(row):
        dif = row['Dif']
        sales = row['Sales']
        actual_added = row['Actual added Current month']
        
        # Convert numerical columns to float if they are strings
        try:
            dif = float(dif) if dif not in ['-', 'No Sales found'] else dif
            sales = float(sales) if sales not in ['-', 'No Sales found'] else sales
            actual_added = float(actual_added) if actual_added not in ['-', 'No Sales found'] else actual_added
        except ValueError:
            pass
        
        if dif == '-':
            return ''
        if row['Delayed Billing'] == 'Delayed Billing':
            return ''
        if (isinstance(dif, (int, float)) and isinstance(sales, (int, float)) and dif > sales) or \
           (isinstance(dif, (int, float)) and dif > 0 and sales == 'No Sales found' and dif == actual_added):
            return "Check Open SO's"
        if (isinstance(dif, (int, float)) and dif != sales and dif != actual_added and 
            sales == 'No Sales found' and actual_added != sales):
            return "Check Device report, or changes to account have been made"
        if isinstance(dif, (int, float)) and dif < 0:
            return 'Cancellations'
        return ''

    df_merged['Notes'] = df_merged.apply(generate_notes, axis=1)

    # Reorder columns to place 'Actual added Current month' after 'Sales' and before 'Delayed Billing'
    column_order = [
        'SabreCode', 'BranchName', 'ItemCode', 'BillingAmount', 'Quantity_Current', 'TotalDue', 
        'Quantity_Previous', 'Dif', 'Sales', 'Actual added Current month', 'Delayed Billing', 'Notes'
    ]
    df_merged = df_merged[column_order]

    # Order the rows
    df_merged = df_merged.sort_values(by=['SabreCode', 'ItemCode', 'BranchName'])

    import logging

    # Configure logging to debug the issue
    logging.basicConfig(filename='process_files_debug.log', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

    try:
        # Ensure the save_path has the correct .xlsx extension
        if not save_path.endswith('.xlsx'):
            save_path += '.xlsx'

        # Convert save_path to an absolute path
        save_path = os.path.abspath(save_path)
        logging.info(f"Saving file to: {save_path}")

        # Save to Excel without formatting
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df_merged.to_excel(writer, index=False, sheet_name='Consolidated Data')
            logging.info("File saved successfully without formatting.")

        # Load the workbook to format it
        workbook = load_workbook(save_path)
        sheet = workbook['Consolidated Data']

        # Define styles
        blue_font = Font(color='0000CC')
        red_font = Font(color='C00000')
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        light_green_fill = PatternFill(start_color='66FF66', end_color='66FF66', fill_type='solid')

        # Find column indices
        col_indices = {cell.value: idx + 1 for idx, cell in enumerate(sheet[1])}
        
        dif_col_index = col_indices.get('Dif')
        sales_col_index = col_indices.get('Sales')
        actual_added_col_index = col_indices.get('Actual added Current month')
        delayed_col_index = col_indices.get('Delayed Billing')

        if dif_col_index is None or sales_col_index is None or actual_added_col_index is None or delayed_col_index is None:
            messagebox.showerror("Error", "One or more required columns not found.")
            return

        for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
            dif_cell = row[dif_col_index - 1]  # -1 for 0-based index
            sales_cell = row[sales_col_index - 1]
            actual_added_cell = row[actual_added_col_index - 1]
            delayed_cell = row[delayed_col_index - 1]

            # Apply blue font to the specified columns
            for cell in [dif_cell, sales_cell, actual_added_cell, delayed_cell]:
                if cell.value not in ['-', 'No Sales found']:
                    cell.font = blue_font

            if isinstance(dif_cell.value, (int, float)):
                if dif_cell.value < 0:
                    dif_cell.font = red_font
                if dif_cell.value == sales_cell.value:
                    dif_cell.fill = green_fill

            if isinstance(sales_cell.value, (int, float)) and isinstance(actual_added_cell.value, (int, float)):
                if sales_cell.value == actual_added_cell.value and sales_cell.value != '-':
                    sales_cell.fill = light_green_fill
                    actual_added_cell.fill = light_green_fill

            if sales_cell.value == 'No Sales found':
                sales_cell.font = red_font

        # Freeze the top row
        sheet.freeze_panes = 'A2'

        # Save the formatted workbook
        workbook.save(save_path)
        logging.info("File saved successfully with formatting.")

    except Exception as e:
        logging.error(f"An error occurred: {e}")
