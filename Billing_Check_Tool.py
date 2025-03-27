import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.dimensions import ColumnDimension
from datetime import datetime

def process_files(prior_csv, current_csv, sales_csv, delayed_csv, client_device_csv, selected_month, selected_year, save_path):
    # Load the CSV files into DataFrames
    df_prior = pd.read_csv(prior_csv)
    df_current = pd.read_csv(current_csv)
    df_sales = pd.read_csv(sales_csv)
    df_delayed = pd.read_csv(delayed_csv)
    df_client_device = pd.read_csv(client_device_csv)

    # Consolidate Current month data
    df_current_consolidated = df_current.groupby(
        ['SabreCode', 'BranchName', 'ItemCode', 'BillingAmount']
    ).agg(
        Quantity_Current=('Quantity', 'sum'),
        TotalDue=('TotalDue', 'sum')
    ).reset_index()

    # Consolidate Prior month data
    df_prior_consolidated = df_prior.groupby(
        ['SabreCode', 'BranchName', 'ItemCode', 'BillingAmount']
    ).agg(
        Quantity_Previous=('Quantity', 'sum')
    ).reset_index()

    # Merge the DataFrames on common columns
    df_merged = pd.merge(
        df_current_consolidated,
        df_prior_consolidated,
        on=['SabreCode', 'BranchName', 'ItemCode', 'BillingAmount'],
        how='left'
    )

    # Calculate the 'Dif' column
    df_merged['Dif'] = df_merged['Quantity_Current'] - df_merged['Quantity_Previous'].fillna(0)

    # Consolidate Sales data
    df_sales_consolidated = df_sales.groupby(
        ['SabreCode', 'TechtoolCode']
    ).agg(
        Sales_Qty=('Qty', 'sum')
    ).reset_index()

    # Merge Sales data into the merged DataFrame
    df_merged['Sales'] = 0  # Initialize the Sales column
    for idx, row in df_merged.iterrows():
        if row['Dif'] > 0:
            sales_data = df_sales_consolidated[
                (df_sales_consolidated['SabreCode'] == row['SabreCode']) &
                (df_sales_consolidated['TechtoolCode'] == row['ItemCode'])  # Assuming 'TechtoolCode' maps to 'ItemCode'
            ]
            if not sales_data.empty:
                df_merged.at[idx, 'Sales'] = sales_data['Sales_Qty'].sum()

    # Update 'Sales' column to 'No Sales found' where 'Dif' > 0 and 'Sales' is 0
    df_merged['Sales'] = df_merged.apply(lambda row: 'No Sales found' if row['Dif'] > 0 and row['Sales'] == 0 else row['Sales'], axis=1)

    # Add 'Delayed Billing' column
    delayed_codes = df_delayed['SabreCode'].unique()
    df_merged['Delayed Billing'] = df_merged['SabreCode'].apply(lambda x: 'Delayed Billing' if x in delayed_codes else '')

    # Add 'Actual added Current month' column
    selected_date = f"{selected_year}-{int(selected_month):02d}"
    df_client_device['ManufactureDate'] = pd.to_datetime(df_client_device['ManufactureDate']).dt.to_period('M')
    
    def calculate_actual_added(row):
        if row['Dif'] > 0:
            actual_added = df_client_device[
                (df_client_device['SabreCode'] == row['SabreCode']) &
                (df_client_device['ItemCode'] == row['ItemCode']) &
                (df_client_device['ManufactureDate'] == selected_date)
            ]['ItemCode'].count()
            return actual_added
        return 0

    df_merged['Actual added Current month'] = df_merged.apply(calculate_actual_added, axis=1)

    # Replace 0 values with '-' in 'Dif', 'Sales', and 'Actual added Current month' columns
    df_merged['Dif'] = df_merged['Dif'].map(lambda x: '-' if x == 0 else x)
    df_merged['Sales'] = df_merged['Sales'].map(lambda x: '-' if x == 0 and x != 'No Sales found' else x)
    df_merged['Actual added Current month'] = df_merged['Actual added Current month'].map(lambda x: '-' if x == 0 else x)

    # Add 'Notes' column and populate it based on conditions
    def generate_notes(row):
        dif = row['Dif']
        sales = row['Sales']
        actual_added = row['Actual added Current month']
        
        # Convert numerical columns to float if they are strings
        try:
            dif = float(dif) if dif not in ['-', 'No Sales found'] else dif
            sales = float(sales) if sales not in ['-', 'No Sales found'] else sales
            actual_added = float(actual_added) if actual_added not in ['-', 'No Sales found'] else actual_added
        except ValueError:
            pass
        
        if dif == '-':
            return ''
        if row['Delayed Billing'] == 'Delayed Billing':
            return ''
        if (isinstance(dif, (int, float)) and isinstance(sales, (int, float)) and dif > sales) or \
           (isinstance(dif, (int, float)) and dif > 0 and sales == 'No Sales found' and dif == actual_added):
            return "Check Open SO's"
        if (isinstance(dif, (int, float)) and dif != sales and dif != actual_added and 
            sales == 'No Sales found' and actual_added != sales):
            return "Check Device report, or changes to account have been made"
        if isinstance(dif, (int, float)) and dif < 0:
            return 'Cancellations'
        return ''

    df_merged['Notes'] = df_merged.apply(generate_notes, axis=1)

    # Reorder columns to place 'Actual added Current month' after 'Sales' and before 'Delayed Billing'
    column_order = [
        'SabreCode', 'BranchName', 'ItemCode', 'BillingAmount', 'Quantity_Current', 'TotalDue', 
        'Quantity_Previous', 'Dif', 'Sales', 'Actual added Current month', 'Delayed Billing', 'Notes'
    ]
    df_merged = df_merged[column_order]

    # Order the rows
    df_merged = df_merged.sort_values(by=['SabreCode', 'ItemCode', 'BranchName'])

    # Save to Excel
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        df_merged.to_excel(writer, index=False, sheet_name='Consolidated Data')

    # Load the workbook to format it
    workbook = load_workbook(save_path)
    sheet = workbook['Consolidated Data']

    # Define styles
    blue_font = Font(color='0000CC')
    red_font = Font(color='C00000')
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    light_green_fill = PatternFill(start_color='66FF66', end_color='66FF66', fill_type='solid')

    # Find column indices
    col_indices = {cell.value: idx + 1 for idx, cell in enumerate(sheet[1])}
    
    dif_col_index = col_indices.get('Dif')
    sales_col_index = col_indices.get('Sales')
    actual_added_col_index = col_indices.get('Actual added Current month')
    delayed_col_index = col_indices.get('Delayed Billing')

    if dif_col_index is None or sales_col_index is None or actual_added_col_index is None or delayed_col_index is None:
        messagebox.showerror("Error", "One or more required columns not found.")
        return

    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
        dif_cell = row[dif_col_index - 1]  # -1 for 0-based index
        sales_cell = row[sales_col_index - 1]
        actual_added_cell = row[actual_added_col_index - 1]
        delayed_cell = row[delayed_col_index - 1]

        # Apply blue font to the specified columns
        for cell in [dif_cell, sales_cell, actual_added_cell, delayed_cell]:
            if cell.value not in ['-', 'No Sales found']:
                cell.font = blue_font

        if isinstance(dif_cell.value, (int, float)):
            if dif_cell.value < 0:
                dif_cell.font = red_font
            if dif_cell.value == sales_cell.value:
                dif_cell.fill = green_fill

        if isinstance(sales_cell.value, (int, float)) and isinstance(actual_added_cell.value, (int, float)):
            if sales_cell.value == actual_added_cell.value and sales_cell.value != '-':
                sales_cell.fill = light_green_fill
                actual_added_cell.fill = light_green_fill

        if sales_cell.value == 'No Sales found':
            sales_cell.font = red_font

    # Freeze the top row
    sheet.freeze_panes = 'A2'

        
    workbook.save(save_path)
    messagebox.showinfo("Success", f"File saved successfully to {save_path}")

# The remaining GUI code is unchanged...
# GUI code
def browse_prior_csv():
    filename = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    if filename:
        entry_prior.delete(0, tk.END)
        entry_prior.insert(0, filename)

def browse_current_csv():
    filename = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    if filename:
        entry_current.delete(0, tk.END)
        entry_current.insert(0, filename)

def browse_sales_csv():
    filename = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    if filename:
        entry_sales.delete(0, tk.END)
        entry_sales.insert(0, filename)

def browse_delayed_csv():
    filename = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    if filename:
        entry_delayed.delete(0, tk.END)
        entry_delayed.insert(0, filename)

def browse_client_device_csv():
    filename = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    if filename:
        entry_client_device.delete(0, tk.END)
        entry_client_device.insert(0, filename)

def browse_save_path():
    filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if filename:
        entry_save.delete(0, tk.END)
        entry_save.insert(0, filename)

def process():
    prior_csv = entry_prior.get()
    current_csv = entry_current.get()
    sales_csv = entry_sales.get()
    delayed_csv = entry_delayed.get()
    client_device_csv = entry_client_device.get()
    selected_month = month_var.get()
    selected_year = year_var.get()
    save_path = entry_save.get()
    
    if not (prior_csv and current_csv and sales_csv and delayed_csv and client_device_csv and selected_month and selected_year and save_path):
        messagebox.showerror("Error", "Please select all files, month, year, and save location.")
        return

    process_files(prior_csv, current_csv, sales_csv, delayed_csv, client_device_csv, selected_month, selected_year, save_path)

# Create the GUI window
root = tk.Tk()
root.title("CSV Processor")

tk.Label(root, text="Select Prior Month CSV:").grid(row=0, column=0, padx=10, pady=5)
entry_prior = tk.Entry(root, width=50)
entry_prior.grid(row=0, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=browse_prior_csv).grid(row=0, column=2, padx=10, pady=5)

tk.Label(root, text="Select Current Month CSV:").grid(row=1, column=0, padx=10, pady=5)
entry_current = tk.Entry(root, width=50)
entry_current.grid(row=1, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=browse_current_csv).grid(row=1, column=2, padx=10, pady=5)

tk.Label(root, text="Select Sales CSV:").grid(row=2, column=0, padx=10, pady=5)
entry_sales = tk.Entry(root, width=50)
entry_sales.grid(row=2, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=browse_sales_csv).grid(row=2, column=2, padx=10, pady=5)

tk.Label(root, text="Select Delayed Billing Client List CSV:").grid(row=3, column=0, padx=10, pady=5)
entry_delayed = tk.Entry(root, width=50)
entry_delayed.grid(row=3, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=browse_delayed_csv).grid(row=3, column=2, padx=10, pady=5)

tk.Label(root, text="Select Client Device Report CSV:").grid(row=4, column=0, padx=10, pady=5)
entry_client_device = tk.Entry(root, width=50)
entry_client_device.grid(row=4, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=browse_client_device_csv).grid(row=4, column=2, padx=10, pady=5)

tk.Label(root, text="Select Month:").grid(row=5, column=0, padx=10, pady=5)
month_var = tk.StringVar(value=datetime.now().strftime('%m'))
month_dropdown = ttk.Combobox(root, textvariable=month_var, values=[f"{i:02d}" for i in range(1, 13)])
month_dropdown.grid(row=5, column=1, padx=10, pady=5)

tk.Label(root, text="Select Year:").grid(row=6, column=0, padx=10, pady=5)
year_var = tk.StringVar(value=datetime.now().strftime('%Y'))
year_dropdown = ttk.Combobox(root, textvariable=year_var, values=[str(i) for i in range(2000, datetime.now().year + 1)])
year_dropdown.grid(row=6, column=1, padx=10, pady=5)

tk.Label(root, text="Save Processed File As:").grid(row=7, column=0, padx=10, pady=5)
entry_save = tk.Entry(root, width=50)
entry_save.grid(row=7, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=browse_save_path).grid(row=7, column=2, padx=10, pady=5)

tk.Button(root, text="Process", command=process).grid(row=8, column=1, padx=10, pady=20)

root.mainloop()